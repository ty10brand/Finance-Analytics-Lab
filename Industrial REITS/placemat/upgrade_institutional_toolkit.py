
"""
upgrade_institutional_toolkit_v2.py

Upgrades an existing REIT placemat workbook (e.g., FR_placemat.xlsx) into an
"institutional v1" toolkit by adding:

- Market_Data: price/returns/shares/market cap (robust market-data fallback)
- Comps: peer snapshot (Yahoo fields; may be missing)
- NAV_Lite: cap-rate scenario NAV shell (input-driven)

Key improvement vs v1:
- Market data uses Yahoo (yfinance) first, then falls back to Stooq.
- If both fail, workbook generation continues (Market_Data shows an error note).

Usage:
  python upgrade_institutional_toolkit_v2.py --in "outputs/models/FR_placemat.xlsx" --out "outputs/models/FR_institutional.xlsx" --user-agent "IndustrialREITPlacemat/1.0 (email@domain.com)"
"""

import os
import time
import re
from io import StringIO
from typing import Optional, List, Dict

import pandas as pd
import requests
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ----------------------------
# CONFIG
# ----------------------------
SEC_FACTS_URL = "https://data.sec.gov/api/xbrl/companyfacts/CIK{cik}.json"

DEFAULT_USER_AGENT = "IndustrialREITPlacemat/1.0 (ty10brand@gmail.com)"

# Peer set for comps (edit as desired)
DEFAULT_PEERS = ["PLD", "STAG", "REXR", "EGP", "TRNO", "LXP", "PLYM", "ILPT", "COLD"]


# ----------------------------
# Styles (simple + clean)
# ----------------------------
FONT_H1 = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
FONT_HDR = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_LBL = Font(name="Calibri", size=11, bold=True)
FONT_VAL = Font(name="Calibri", size=11)

FILL_DARK = PatternFill("solid", fgColor="1F4E79")
FILL_HDR = PatternFill("solid", fgColor="0B2F4A")
FILL_INPUT = PatternFill("solid", fgColor="FFF2CC")

ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_R = Alignment(horizontal="right", vertical="center")
ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)

thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

FMT_USD2 = '$#,##0.00;[Red]($#,##0.00);"-"'
FMT_USD0 = '$#,##0;[Red]($#,##0);"-"'
FMT_PCT1 = '0.0%;[Red](0.0%);"-"'
FMT_X1 = '0.0"x"'


def header_bar(ws, row, col_start, col_end, text):
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=text)
    c.font = FONT_H1
    c.fill = FILL_DARK
    c.alignment = ALIGN_L
    for col in range(col_start, col_end + 1):
        ws.cell(row=row, column=col).fill = FILL_DARK


def table_header(ws, row, cols):
    for i, name in enumerate(cols, start=1):
        cell = ws.cell(row=row, column=i, value=name)
        cell.font = FONT_HDR
        cell.fill = FILL_HDR
        cell.alignment = ALIGN_C
        cell.border = BORDER


def set_widths(ws, widths: Dict[str, float]):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def ensure_sheet(wb, name: str):
    if name in wb.sheetnames:
        ws = wb[name]
        ws.delete_rows(1, ws.max_row)
        return ws
    return wb.create_sheet(name)


# ----------------------------
# SEC helpers
# ----------------------------
def sec_get_companyfacts(cik: str, user_agent: str) -> dict:
    url = SEC_FACTS_URL.format(cik=cik)
    headers = {
        "User-Agent": user_agent,
        "Accept-Encoding": "gzip, deflate",
        "Host": "data.sec.gov",
    }
    r = requests.get(url, headers=headers, timeout=30)
    r.raise_for_status()
    return r.json()


def extract_latest_shares(companyfacts: dict) -> Optional[float]:
    """
    Best-effort shares outstanding.
    Prefer us-gaap:CommonStockSharesOutstanding if available.
    """
    facts = companyfacts.get("facts", {}).get("us-gaap", {})
    tag = "CommonStockSharesOutstanding"
    if tag not in facts:
        return None

    units = facts[tag].get("units", {})
    share_unit = None
    for k in units.keys():
        if "shares" in k.lower():
            share_unit = k
            break
    if not share_unit:
        return None

    df = pd.DataFrame(units[share_unit])
    if "end" not in df.columns or "val" not in df.columns:
        return None

    df["end"] = pd.to_datetime(df["end"], errors="coerce")
    df = df.dropna(subset=["end", "val"]).sort_values("end")
    if df.empty:
        return None

    return float(df.iloc[-1]["val"])


# ----------------------------
# Market helpers (Yahoo + Stooq fallback)
# ----------------------------
def get_price_history(ticker: str, start="2010-01-01") -> pd.Series:
    """
    Tries Yahoo (yfinance) first; falls back to Stooq daily close.
    NOTE: Stooq close is not dividend-adjusted; Yahoo auto_adjust is.
    """
    start_dt = pd.to_datetime(start)

    # 1) Yahoo via yfinance
    try:
        df = yf.download(ticker, start=start, auto_adjust=True, progress=False)
        if df is not None and not df.empty and "Close" in df.columns:
            s = df["Close"].dropna()
            if not s.empty:
                s.name = ticker
                return s
    except Exception:
        pass

    # 2) Stooq fallback (daily close)
    try:
        sym = f"{ticker.lower()}.us"
        url = f"https://stooq.com/q/d/l/?s={sym}&i=d"
        r = requests.get(url, timeout=30)
        r.raise_for_status()
        txt = r.text.strip()

        if "Date" not in txt or "Close" not in txt:
            raise RuntimeError(f"Stooq response not recognized for {ticker}")

        df2 = pd.read_csv(StringIO(txt))
        df2["Date"] = pd.to_datetime(df2["Date"], errors="coerce")
        df2 = df2.dropna(subset=["Date"]).set_index("Date").sort_index()
        df2 = df2[df2.index >= start_dt]

        if "Close" not in df2.columns:
            raise RuntimeError(f"Stooq missing Close for {ticker}")

        s = pd.to_numeric(df2["Close"], errors="coerce").dropna()
        if s.empty:
            raise RuntimeError(f"Stooq returned no prices for {ticker}")

        s.name = ticker
        return s
    except Exception:
        pass

    raise RuntimeError(f"No price history for {ticker} (Yahoo + Stooq failed)")


def total_return_from_to(price: pd.Series, start_date: pd.Timestamp, end_date: pd.Timestamp) -> Optional[float]:
    p = price.dropna()
    if p.empty:
        return None
    start_idx = p.index[p.index.get_indexer([start_date], method="nearest")[0]]
    end_idx = p.index[p.index.get_indexer([end_date], method="nearest")[0]]
    return float(p.loc[end_idx] / p.loc[start_idx] - 1.0)


def get_yf_info(ticker: str) -> dict:
    try:
        return yf.Ticker(ticker).info or {}
    except Exception:
        return {}


# ----------------------------
# Sheet writers
# ----------------------------
def write_market_data_sheet(wb, ticker: str, shares: Optional[float]):
    ws = ensure_sheet(wb, "Market_Data")
    ws.sheet_view.showGridLines = False
    set_widths(ws, {"A": 28, "B": 22, "C": 45})

    header_bar(ws, 1, 1, 3, f"Market Data — {ticker}")

    try:
        price = get_price_history(ticker, start="2010-01-01")
        last_price = float(price.iloc[-1])
        asof = price.index[-1]

        r1 = total_return_from_to(price, asof - pd.Timedelta(days=365), asof)
        r3 = total_return_from_to(price, asof - pd.Timedelta(days=365 * 3), asof)
        r5 = total_return_from_to(price, asof - pd.Timedelta(days=365 * 5), asof)

        info = get_yf_info(ticker)
        yf_mktcap = info.get("marketCap", None)
        mktcap = float(yf_mktcap) if yf_mktcap else (last_price * shares if shares else None)

        rows = [
            ("As of", asof.strftime("%Y-%m-%d"), ""),
            ("Last price ($)", last_price, "Yahoo adjusted close or Stooq close (fallback)"),
            ("Shares outstanding (mm)", (shares / 1e6) if shares else None, "Best-effort from SEC XBRL"),
            ("Market cap ($mm)", (mktcap / 1e6) if mktcap else None, "Yahoo marketCap else price*shares"),
            ("Total return 1Y", r1, "Using available price series"),
            ("Total return 3Y", r3, "Using available price series"),
            ("Total return 5Y", r5, "Using available price series"),
        ]
    except Exception as e:
        # Graceful: workbook still builds
        ws["A3"] = "Market data error"
        ws["A3"].font = FONT_LBL
        ws["A3"].alignment = ALIGN_L
        ws["B3"] = str(e)
        ws["B3"].alignment = ALIGN_L
        ws["B3"].font = FONT_VAL
        return

    # header row
    ws["A3"] = "Metric"
    ws["B3"] = "Value"
    ws["C3"] = "Notes"
    for c in ["A3", "B3", "C3"]:
        ws[c].font = FONT_HDR
        ws[c].fill = FILL_HDR
        ws[c].alignment = ALIGN_C
        ws[c].border = BORDER

    r0 = 4
    for i, (m, v, note) in enumerate(rows):
        r = r0 + i

        ws[f"A{r}"] = m
        ws[f"A{r}"].font = FONT_LBL
        ws[f"A{r}"].alignment = ALIGN_L
        ws[f"A{r}"].border = BORDER

        ws[f"B{r}"] = v
        ws[f"B{r}"].font = FONT_VAL
        ws[f"B{r}"].alignment = ALIGN_R
        ws[f"B{r}"].border = BORDER

        ws[f"C{r}"] = note
        ws[f"C{r}"].font = FONT_VAL
        ws[f"C{r}"].alignment = ALIGN_L
        ws[f"C{r}"].border = BORDER

        if "price" in m.lower():
            ws[f"B{r}"].number_format = FMT_USD2
        elif "shares" in m.lower() or "cap" in m.lower():
            ws[f"B{r}"].number_format = FMT_USD0
        elif "return" in m.lower():
            ws[f"B{r}"].number_format = FMT_PCT1


def write_comps_sheet(wb, universe: List[str]):
    ws = ensure_sheet(wb, "Comps")
    ws.sheet_view.showGridLines = False
    set_widths(ws, {"A": 10, "B": 14, "C": 18, "D": 18, "E": 16, "F": 16, "G": 40})

    header_bar(ws, 1, 1, 7, "Comps (Public Snapshot) — Yahoo fields where available")
    cols = ["Ticker", "Price", "Market Cap ($mm)", "Enterprise Value ($mm)", "Shares (mm)", "Dividend Yield", "Notes"]
    table_header(ws, 3, cols)

    start_row = 4
    for i, t in enumerate(universe):
        r = start_row + i
        info = get_yf_info(t)

        price = info.get("currentPrice", None)
        mktcap = info.get("marketCap", None)
        ev = info.get("enterpriseValue", None)
        div_yld = info.get("dividendYield", None)
        shares = info.get("sharesOutstanding", None)

        note = "Yahoo snapshot (some fields may be missing)"
        ws.cell(r, 1, t).border = BORDER
        ws.cell(r, 1).alignment = ALIGN_C

        ws.cell(r, 2, price).number_format = FMT_USD2
        ws.cell(r, 3, (mktcap / 1e6 if mktcap else None)).number_format = FMT_USD0
        ws.cell(r, 4, (ev / 1e6 if ev else None)).number_format = FMT_USD0
        ws.cell(r, 5, (shares / 1e6 if shares else None)).number_format = FMT_USD0
        ws.cell(r, 6, div_yld).number_format = FMT_PCT1
        ws.cell(r, 7, note)

        for c in range(2, 8):
            ws.cell(r, c).border = BORDER
            ws.cell(r, c).alignment = ALIGN_R if c in (2, 3, 4, 5, 6) else ALIGN_L


def write_nav_lite_sheet(wb, base_ticker: str):
    ws = ensure_sheet(wb, "NAV_Lite")
    ws.sheet_view.showGridLines = False
    set_widths(ws, {"A": 30, "B": 18, "C": 34, "D": 18})

    header_bar(ws, 1, 1, 4, f"NAV (Cap Rate Scenarios) — {base_ticker}")

    ws["A3"] = "Inputs (yellow)"
    ws["A3"].font = FONT_LBL

    inputs = [
        ("Stabilized NOI ($mm)", None, "Input from supplemental / NOI build"),
        ("Net Debt ($mm)", "=Valuation!C5", "Link if you fill Net Debt on Valuation tab"),
        ("Shares (mm)", "=Market_Data!B6", "Link from Market_Data"),
        ("Cap rate (Base)", 0.050, "Example 5.0%"),
        ("Cap rate (Bull)", 0.045, "Example 4.5%"),
        ("Cap rate (Bear)", 0.055, "Example 5.5%"),
    ]

    for i, (lab, val, note) in enumerate(inputs):
        r = 4 + i
        ws[f"A{r}"] = lab
        ws[f"A{r}"].font = FONT_LBL
        ws[f"A{r}"].border = BORDER
        ws[f"A{r}"].alignment = ALIGN_L

        ws[f"B{r}"] = val
        ws[f"B{r}"].border = BORDER
        ws[f"B{r}"].alignment = ALIGN_R
        ws[f"B{r}"].fill = FILL_INPUT
        if "cap rate" in lab.lower():
            ws[f"B{r}"].number_format = FMT_PCT1
        else:
            ws[f"B{r}"].number_format = FMT_USD0

        ws[f"C{r}"] = note
        ws[f"C{r}"].border = BORDER
        ws[f"C{r}"].alignment = ALIGN_L

    ws["A11"] = "Outputs (Base scenario)"
    ws["A11"].font = FONT_LBL

    ws["A12"] = "Implied Property Value ($mm)"
    ws["A13"] = "NAV Equity Value ($mm)"
    ws["A14"] = "NAV / Share ($)"
    ws["A15"] = "Premium/(Discount) to NAV"
    for r in range(12, 16):
        ws[f"A{r}"].font = FONT_LBL
        ws[f"A{r}"].border = BORDER
        ws[f"A{r}"].alignment = ALIGN_L

    ws["B12"] = '=IF(OR(B4="",B7=""),"",B4/B7)'
    ws["B13"] = '=IF(OR(B12="",B5=""),"",B12-B5)'
    ws["B14"] = '=IF(OR(B13="",B6=""),"",B13/B6)'
    ws["B15"] = '=IF(OR(B14="",Market_Data!B5=""),"",Market_Data!B5/B14-1)'

    for addr, fmt in [("B12", FMT_USD0), ("B13", FMT_USD0), ("B14", FMT_USD2), ("B15", FMT_PCT1)]:
        ws[addr].number_format = fmt
        ws[addr].border = BORDER
        ws[addr].alignment = ALIGN_R

    ws["C12"] = "Base cap rate"
    ws["C12"].alignment = ALIGN_L
    ws["C12"].border = BORDER

    ws["D11"] = "Bull/Bear NAV per share ($)"
    ws["D11"].font = FONT_LBL

    ws["D12"] = '=IF(OR(B4="",B8=""),"", (B4/B8 - B5)/B6)'  # bull
    ws["D13"] = '=IF(OR(B4="",B9=""),"", (B4/B9 - B5)/B6)'  # bear
    ws["D12"].number_format = FMT_USD2
    ws["D13"].number_format = FMT_USD2
    ws["D12"].border = BORDER
    ws["D13"].border = BORDER
    ws["D12"].alignment = ALIGN_R
    ws["D13"].alignment = ALIGN_R

    ws["C16"] = "Bull (cap rate)"
    ws["C17"] = "Bear (cap rate)"
    ws["D16"] = "=B8"
    ws["D17"] = "=B9"
    ws["D16"].number_format = FMT_PCT1
    ws["D17"].number_format = FMT_PCT1
    for a in ["C16", "C17", "D16", "D17"]:
        ws[a].border = BORDER
        ws[a].alignment = ALIGN_L if a.startswith("C") else ALIGN_R


# ----------------------------
# Upgrade runner
# ----------------------------
def upgrade_workbook(
    placemat_path: str,
    out_path: str,
    user_agent: str = DEFAULT_USER_AGENT,
    peers: Optional[List[str]] = None,
):
    if peers is None:
        peers = DEFAULT_PEERS

    wb = load_workbook(placemat_path)

    # Pull ticker & cik from Inputs (your template layout)
    if "Inputs" not in wb.sheetnames:
        raise RuntimeError("Workbook missing 'Inputs' sheet.")
    ticker = wb["Inputs"]["C4"].value
    cik = wb["Inputs"]["C5"].value

    if not ticker or not cik:
        raise RuntimeError("Inputs sheet must have ticker in C4 and cik in C5.")

    ticker = str(ticker).upper().strip()
    cik = str(cik).zfill(10)

    # SEC shares best-effort
    facts = sec_get_companyfacts(cik, user_agent=user_agent)
    shares = extract_latest_shares(facts)

    # Create/refresh tabs
    write_market_data_sheet(wb, ticker=ticker, shares=shares)

    # Comps universe includes the base ticker first
    comps_universe = [ticker] + [t.upper() for t in peers if t.upper() != ticker]
    write_comps_sheet(wb, comps_universe)

    write_nav_lite_sheet(wb, base_ticker=ticker)

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    print(f"[OK] Upgraded workbook written: {out_path}")


def main():
    import argparse

    p = argparse.ArgumentParser()
    p.add_argument("--in", dest="inp", required=True, help="Input placemat xlsx (e.g., outputs/models/FR_placemat.xlsx)")
    p.add_argument("--out", dest="outp", required=True, help="Output upgraded workbook xlsx")
    p.add_argument("--user-agent", default=DEFAULT_USER_AGENT, help="SEC-compliant User-Agent")
    p.add_argument("--peers", nargs="*", default=DEFAULT_PEERS, help="Peer tickers for comps")

    args = p.parse_args()

    upgrade_workbook(
        placemat_path=args.inp,
        out_path=args.outp,
        user_agent=args.user_agent,
        peers=[t.upper() for t in args.peers],
    )


if __name__ == "__main__":
    main()




