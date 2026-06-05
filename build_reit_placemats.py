

import os
import json
import re
import time
import requests
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

TEMPLATE_PATH = "templates/industrial_reit_placemat_template.xlsx"
OUT_DIR = "outputs/models"
CIK_MAP_PATH = "data/cik/ticker_cik_map.json"

SEC_FACTS_URL = "https://data.sec.gov/api/xbrl/companyfacts/CIK{cik}.json"

# Common GAAP tags (not perfect for every REIT, but good v1 coverage)
GAAP_TAGS = {
    # Income statement (annual)
    "Total Revenue": ["Revenues", "SalesRevenueNet"],
    "Operating income": ["OperatingIncomeLoss"],
    "Interest expense": ["InterestExpense"],
    "Net income": ["NetIncomeLoss"],

    # Balance sheet (FY end)
    "Total assets": ["Assets"],
    "Total liabilities": ["Liabilities"],
    "Total debt": ["LongTermDebt", "LongTermDebtNoncurrent", "DebtLongtermAndShorttermCombinedAmount"],
    "Equity": ["StockholdersEquity", "StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest"],

    # Cash flow (annual)
    "Net cash from operating activities": ["NetCashProvidedByUsedInOperatingActivities"],
    "Net cash from investing activities": ["NetCashProvidedByUsedInInvestingActivities"],
    "Net cash from financing activities": ["NetCashProvidedByUsedInFinancingActivities"],
    # Capex (best-effort; REITs sometimes report multiple variants)
    "Capex (total)": ["PaymentsToAcquireRealEstate", "PaymentsToAcquirePropertyPlantAndEquipment"],
}

def norm_ticker(t: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", t.upper().strip())

def sec_get_json(url: str, user_agent: str) -> dict:
    headers = {
        "User-Agent": user_agent,
        "Accept-Encoding": "gzip, deflate",
        "Host": "data.sec.gov",
    }
    r = requests.get(url, headers=headers, timeout=30)
    r.raise_for_status()
    return r.json()

def load_cik_map() -> Dict[str, str]:
    with open(CIK_MAP_PATH, "r", encoding="utf-8") as f:
        return json.load(f)

def pick_annual_usd_facts(facts_json: dict, tags: List[str]) -> Optional[pd.DataFrame]:
    """
    Return a DataFrame of annual USD values for the first GAAP tag that exists.
    Filters to FY (annual) rows where available and prefers 10-K.
    """
    facts = facts_json.get("facts", {}).get("us-gaap", {})
    for tag in tags:
        if tag not in facts:
            continue
        units = facts[tag].get("units", {})
        # Prefer USD
        if "USD" not in units:
            continue
        rows = units["USD"]
        df = pd.DataFrame(rows)

        # Try to keep annual points (FY). Many facts include 'fy' and 'fp'.
        # Prefer fp == 'FY' and form == '10-K' when present.
        if "fp" in df.columns:
            df = df[df["fp"] == "FY"]

        if "form" in df.columns:
            df_10k = df[df["form"] == "10-K"]
            if not df_10k.empty:
                df = df_10k

        # Keep essentials
        for col in ["end", "fy", "val"]:
            if col not in df.columns:
                return None

        df["end"] = pd.to_datetime(df["end"], errors="coerce")
        df = df.dropna(subset=["end", "val", "fy"]).copy()
        df["fy"] = df["fy"].astype(int)

        # If duplicates per FY, keep the latest 'end'
        df = df.sort_values(["fy", "end"]).drop_duplicates(subset=["fy"], keep="last")

        return df[["fy", "end", "val"]].rename(columns={"val": tag})
    return None

def get_last_n_fys(facts_json: dict, n: int = 5) -> List[int]:
    # Use Assets as a reliable anchor
    anchor = pick_annual_usd_facts(facts_json, ["Assets"])
    if anchor is None or anchor.empty:
        return []
    fys = sorted(anchor["fy"].unique().tolist())
    return fys[-n:]

def fill_template(
    ticker: str,
    cik: str,
    company_name: str,
    facts_json: dict,
    out_path: str
):
    wb = load_workbook(TEMPLATE_PATH)
    ws_inputs = wb["Inputs"]
    ws_fin = wb["Financials_Annual"]

    # Determine last 5 fiscal years available
    fys = get_last_n_fys(facts_json, n=5)
    if len(fys) < 3:
        raise RuntimeError(f"Not enough FY data found for {ticker} (CIK {cik}). Found fys={fys}")

    start_fy, end_fy = fys[0], fys[-1]

    # Fill Inputs sheet (matches template layout we created)
    ws_inputs["C3"].value = company_name
    ws_inputs["C4"].value = ticker
    ws_inputs["C5"].value = cik
    ws_inputs["C9"].value = start_fy
    ws_inputs["C10"].value = end_fy

    # Map FY -> column in Financials_Annual
    year_cols = ["C", "D", "E", "F", "G"]
    fy_to_col = {fy: year_cols[i] for i, fy in enumerate(fys)}

    # Helper to write a line item into Financials_Annual by matching label in col B
    def write_financial_line(label: str, values_by_fy: Dict[int, float]):
        # Find the label row in column B
        for row in range(1, ws_fin.max_row + 1):
            if (ws_fin.cell(row=row, column=2).value or "").strip() == label:
                for fy, val in values_by_fy.items():
                    col = fy_to_col.get(fy)
                    if col is None:
                        continue
                    ws_fin[f"{col}{row}"].value = float(val)
                return
        # If label doesn't exist in template, ignore (template has some placeholders)
        return

    # Pull each metric and write into template
    for line_label, tags in GAAP_TAGS.items():
        df = pick_annual_usd_facts(facts_json, tags)
        if df is None or df.empty:
            continue

        # df column name is the actual GAAP tag used; normalize to values_by_fy
        used_tag = [c for c in df.columns if c not in ("fy", "end")][0]
        values_by_fy = dict(zip(df["fy"].tolist(), df[used_tag].tolist()))

        write_financial_line(line_label, values_by_fy)

    # Save
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)

def main():
    import argparse
    p = argparse.ArgumentParser()
    p.add_argument("--user-agent", required=True, help="Name/Project (email@domain.com)")
    p.add_argument("--tickers", nargs="*", default=None, help="Optional list of tickers. If omitted, uses config/tickers.yaml reits.")
    args = p.parse_args()

    # Load tickers list
    tickers = args.tickers
    if tickers is None or len(tickers) == 0:
        # Pull from your existing config/tickers.yaml to match your universe
        import yaml
        with open("config/tickers.yaml", "r") as f:
            cfg = yaml.safe_load(f)
        tickers = cfg.get("reits", [])
    tickers = [norm_ticker(t) for t in tickers]

    cik_map = load_cik_map()

    for t in tickers:
        if t not in cik_map:
            print(f"[SKIP] No CIK found for {t} in local map.")
            continue

        cik = cik_map[t]
        url = SEC_FACTS_URL.format(cik=cik)
        facts = sec_get_json(url, args.user_agent)

        company_name = facts.get("entityName", t)
        out_path = os.path.join(OUT_DIR, f"{t}_placemat.xlsx")

        fill_template(ticker=t, cik=cik, company_name=company_name, facts_json=facts, out_path=out_path)

        print(f"[OK] Built {out_path}")
        # Be polite to SEC
        time.sleep(0.25)

if __name__ == "__main__":
    main()

